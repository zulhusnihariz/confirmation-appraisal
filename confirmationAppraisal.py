# ---------------------------------------------------------------------------- #
#                               Import Libraries                               #
# ---------------------------------------------------------------------------- #
import ast
from datetime import datetime
import os
import pandas as pd
import tkinter as tk
import win32com.client as client
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException
from textwrap import dedent
from tkinter import Button, END, Entry, filedialog, Frame, Label, LabelFrame, messagebox, PhotoImage, Scrollbar, StringVar, Text, Toplevel,ttk

# ---------------------------------------------------------------------------- #
#                                  Main Frame                                  #
# ---------------------------------------------------------------------------- #
class Application:
    def __init__(self, root):
        self.root = root
        self.root.geometry('1100x430')
        self.root.title("Confirmation Appraisal")
        self.root.iconbitmap(True, 'rawr.ico')
        self.root.resizable(0,0)
        self.create_menubar()
        self.deleteName = []

        self.wrapper = LabelFrame(root, text = "Working File Name")
        self.wrapper.grid(row = 0, column = 0, padx = 10, pady = 5,  sticky = 'w')

        # --------------------------- List of textvariable --------------------------- #
        self.fileName = StringVar()
        self.emailSubject = StringVar()
        self.emailBody = StringVar()
        self.defCC = StringVar()
        
        self.savedPath()
        self.savedSubject()
        self.savedBody()

        self.fileValidity()
        print('file is valid')
        self.setupList()
        self.treeView()
        self.entry()
        self.button()

    def append_deleteName(self, new_value):
        self.deleteName.append(new_value)

    def savedPath(self):
        with open('filePath.txt', 'r') as f:
            self.askName = ast.literal_eval(f.read()) # Literal String
            self.fileName.set(os.path.split(self.askName)[1]) # Get File Name to be Displayed
            return self.fileName

    def savedBody(self):
        with open('outlookBody.txt', 'r') as f:
            self.emailBody = f.read()
            # print(self.emailBody)
            return self.emailBody

    def savedSubject(self):
        with open('outlookSubject.txt', 'r') as f:
            self.emailSubject = f.read()
            # print(self.emailSubject)
            return self.emailSubject

# ---------------------------------------------------------------------------- #
#       Set and Save Working File for Future Use until Explicitly Changed      #
# ---------------------------------------------------------------------------- #
    def browsePath(self):
        self.askName = filedialog.askopenfilename(initialdir = os.getcwd(), title = 'Select a file', filetypes=[("Excel files", ".xlsx .xls")])
        self.fileName.set(os.path.split(self.askName)[1])

        self.fileValidity()
        self.setupList()
        self.deleteName.clear()
        self.savedSubject()
        self.savedBody()
        with open('filePath.txt', 'w') as f:
            f.write(f"r'{self.askName}'")
        try:
            if self.myTree is None:
                pass
            else:
                self.clearTree()
        except:
            pass
        return self.fileName, self.askName, self.emailBody  

# ---------------------------------------------------------------------------- #
#       Add End of Probation Column to Staff Listing (to be run monthly)       #
# ---------------------------------------------------------------------------- #
    def addEndProbation(self):
    # ------------------------ Load File and Sheet by Name ----------------------- #
        wb = load_workbook(self.askName)
        ws = wb.active

        if len(wb.sheetnames) > 1:
            messagebox.showinfo(title="Error", message=f"Please leave only one sheet in the listing and ensure no hidden column & sheets")
            self.closeWindow()

    # ---------------- Insert New Column for End of Probation Date --------------- #
        for val in ws[1]:
            if val.value == 'DtJoined':
                cellColumn = val.column + 1
                cellRow = val.row +1
                dtJoinedCoordinate = val.coordinate
                endProbCoordinate = ws.cell(row = cellRow, column = cellColumn).coordinate
                # print(dtJoinedCoordinate) # column name coordinate
                # print(endProbCoordinate) # column name coordinate
                ws.insert_cols(cellColumn)

        # ------------- Select The New Column, Define Styles(Font & Border) ------------- #
        wb.save(self.askName)
        endProb = ws[str(endProbCoordinate[0])]
        K2 = ws[f'{dtJoinedCoordinate[0]}2']
        pttrnfll = PatternFill(start_color = 'FFC83D', end_color = 'FFC83D', fill_type= 'solid')
        # ----- Copy Styles, Values (Formulas), & Number Formatting for Each Cell ----- #
        for val in endProb:
            val._style = copy(K2._style)
            val.value = f'=EDATE({dtJoinedCoordinate[0]}{val.row},6)'
        # ----------------------- Formatting for Column Header ----------------------- #
        endProbHeader = ws[f'{endProbCoordinate[0]}1']
        endProbHeader.value = 'endProbation'
        endProbHeader.fill = pttrnfll
        endProbHeader.font = Font(name ='Trebuchet MS', size = '9', bold=True)
        ws.column_dimensions[f'{endProbCoordinate[0]}'].width = 15
        # ----------------------- Save Changes Made to Workbook ---------------------- #
        # print(ws['I1'].value)
        # print(ws['J1'].value)
        wb.save(self.askName)
        # ---------------------- Read & Save File Using pywin32 ---------------------- #
        '''This step is necessary; for some reason, pandas wouldn't read "endProbation" column (read values as NaN). 
        So this additional step open the workbook and save it using pywin32.'''
        print('Done!')
        excel = client.Dispatch('Excel.Application')
        workbook = excel.Workbooks.Open(self.askName)
        workbook.Save()
        workbook.Close()
        excel.Quit()
        return self.askName

    def closeWindow(self):
        self.root.destroy()

    def fileValidity(self):
        if self.askName == "":# If DtJoined column does not exists, pop up path filedialog
            # print("first if block")
            result = messagebox.askquestion("No file selected", "Select a file?", icon='warning')
            if result == 'yes':
                self.browsePath() # if no path set during launch; this will pop out
            else:
                self.closeWindow()
        return self.askName

    def dataValidity(self):
        master = pd.read_excel(self.askName, engine = 'openpyxl')
        master.columns = map(str.lower, master.columns)
        validDict = {'pers.no.':'Pers.No.', 'name':'Name', 'email':'Email', 'gender':'Gender','dtjoined':'DtJoined'}

        for key, value in validDict.items():
            if key not in (master.columns): # or 'DtJoined' not in list(master.columns) 
                messagebox.showinfo(title="Error", message=f"Column '{value}' not found. Please select a valid staff listing")
                self.browsePath()
                return
            else:
                pass

# ---------------------------------------------------------------------------- #
#      Setup variables from Excel file; Use to display column in treeview      #
# ---------------------------------------------------------------------------- #
    def setupList(self):
        # ------------------------------ Read Excel File ----------------------------- #
        self.dataValidity()
        master = pd.read_excel(self.askName, engine = 'openpyxl')
        master.columns = map(str.lower, master.columns)

        if 'endprobation' in list(master.columns): # If DtJoined column does not exists, pop up path filedialog
            pass
        else:
            print('no endprobation block')
            self.addEndProbation()
            print('it should exists now')

        pd.set_option('display.max_columns', 1000)
        pd.set_option('display.max_row', 1000)

        master = pd.read_excel(self.askName, engine = 'openpyxl')
        master['DtJoined'] = pd.to_datetime(master['DtJoined'],errors='coerce')

        caseInsensitive = master
        caseInsensitive.columns = map(str.lower, caseInsensitive.columns)
        # print(list(master.columns))

        try:
            self.df = master.loc[:, ('pers.no.', 'name','department','email', 'gender', 'dtjoined','endprobation')]
        except: #if 'department' or 'dept' not in caseInsensitive.columns:
            self.df = master.loc[:, ('pers.no.', 'name','org. unit','email', 'gender', 'dtjoined','endprobation')]
        
        self.df.sort_values('endprobation', ascending = False, inplace = True)
        self.df_col = list(self.df.columns.values)

        return self.df_col
# ---------------------------------------------------------------------------- #
#                                   TreeView                                   #
# ---------------------------------------------------------------------------- #
    def treeView(self):
        self.treeWrapper = LabelFrame(root)
        self.treeWrapper.grid(row = 1, column = 0, padx = 10, pady = 5,  sticky = 'w')

        self.treeFrame = LabelFrame(root, text = 'List of Staff')
        self.treeFrame.grid(row = 2, column = 0, pady = 10, padx = 5)
        
        self.treeScroll =  Scrollbar(self.treeFrame)
        self.treeScroll.pack(side = 'right', fill = 'y')
        
        self.myTree = ttk.Treeview(self.treeFrame, yscrollcommand = self.treeScroll.set)
        self.myTree.pack(pady = 10, padx = 10)
        
        self.treeScroll.config(command = self.myTree.yview)
        
        self.myTree['columns'] = (self.df_col)

        for x in range(len(self.df_col)):
            self.myTree.column(self.df_col[x], width=100)
            self.myTree.heading(self.df_col[x], text=self.df_col[x])

        # # ------------------------------- Format Column ------------------------------ #
        self.myTree.column('#0', width = 50, minwidth = 25, anchor='center', stretch = False)
        self.myTree.heading("#0", text="No.", anchor='center')
       
        self.myTree.column('pers.no.', anchor = 'center', width = 70)
        self.myTree.heading('pers.no.', text="Pers.No.", anchor='center')
       
        self.myTree.column('name', anchor = 'w', width = 220)
        self.myTree.heading('name', text="Name", anchor='center')
        
        self.myTree.column('email', anchor = 'w', width = 250)
        self.myTree.heading('email', text="Email", anchor='center')

        self.myTree.column('gender', anchor = 'w', width = 70)
        self.myTree.heading('gender', text="Gender", anchor='center')

        if 'department' not in self.df_col:
            try:
                self.myTree.column('dept', anchor = 'w', width = 180)
                self.myTree.heading("dept", text="Department", anchor='center')
            except:
                self.myTree.column('org. unit', anchor = 'w', width = 180)
                self.myTree.heading("org. unit", text="Department", anchor='center')
        else:
            self.myTree.column('department', anchor = 'w', width = 180)
            self.myTree.heading("department", text="Department", anchor='center')

        self.myTree.column('dtjoined', anchor = 'center', width = 100)
        self.myTree.heading('dtjoined', text="DtJoined", anchor='center')

        self.myTree.column('endprobation', anchor = 'center', width = 100)
        self.myTree.heading('endprobation', text="endProbation", anchor='center')

        self.root.bind('<Return>', self.parse)

# ---------------------------------------------------------------------------- #
#                                Create Menu Bar                               #
# ---------------------------------------------------------------------------- #
    def create_menubar(self):
        # create the menubar
        self.menubar = tk.Menu(self.root)
        self.root.configure(menu=self.menubar)
        # File menu
        fileMenu = tk.Menu(self.menubar, tearoff = 0)
        self.menubar.add_cascade(label="File", menu=fileMenu)
        fileMenu.add_command(label="Open File", command=self.browsePath)

# ---------------------------------------------------------------------------- #
#           Verify whether input in number of days entry are integer           #
# ---------------------------------------------------------------------------- #
    def number(self):
        try:
            self.numberOfDays = int(self.daysEntry.get())
            print(self.numberOfDays)
            return self.numberOfDays
        except ValueError:
            self.numberOfDays = 0
            self.daysEntry.delete(0, END)
            messagebox.showinfo(title="Error", message="Invalid data type. Number only")
            
# ---------------------------------------------------------------------------- #
#                 Show staff listing whenever button is clicked                #
# ---------------------------------------------------------------------------- #
    def showList(self):
        if self.askName == "":
            messagebox.showinfo(title="Error", message="No file was selected")
            return
        else:
            pass
        # ----------------------- Set Today's Date as Variable ----------------------- #
        current_datetime = datetime.now()
        dateToday = current_datetime.strftime('%Y-%m-%d')

        self.df.loc[:,'Today'] = dateToday
        self.df['Today'] = self.df['Today'].apply(pd.to_datetime)
        a = self.df['endprobation'] - pd.Timedelta(self.numberOfDays, unit='d') 
        b = self.df['endprobation']

        self.end = self.df[self.df['Today'].between(a, b, inclusive=False)]

        self.end.reset_index(inplace = True, drop = True)
        self.end.index = self.end.index + 1

        self.end.loc[:,'dtjoined'] = self.end['dtjoined'].dt.strftime('%d-%m-%Y')
        self.end.loc[:,'endprobation'] = self.end['endprobation'].dt.strftime('%d-%m-%Y')
        
        self.rowLabels = self.end.index.tolist() # Set index for treeview

        for i in range(len(self.end)):# insert index into treeview
            self.myTree.insert('', i, text=self.rowLabels[i], values= self.end.iloc[i,:].tolist())
             

        self.myTree.column('#0', width = 50, minwidth = 25, anchor='center', stretch = False) # reformat treeview
        self.myTree.heading("#0", text="No.", anchor='center')

        for info in range(len(self.rowLabels)):
            self.pfNumber = self.end.loc[self.end.index[info],'pers.no.']
            self.gender = self.end.loc[self.end.index[info],'gender']
            self.name = self.end.loc[self.end.index[info],'name']
            self.email = self.end.loc[self.end.index[info],'email']
            self.dueDate = self.end.loc[self.end.index[info], 'endprobation']

# ---------------------------------------------------------------------------- #
#              Clear TreeView whenever 'Filter' button is clicked              #
# ---------------------------------------------------------------------------- #
    def clearTree(self):
        for i in self.myTree.get_children():
            self.myTree.delete(i)
# ---------------------------------------------------------------------------- #
#                       Delete selected row in Tree View                       #
# ---------------------------------------------------------------------------- #
    def removeMany(self):
        x = self.myTree.selection() #get rows value in form of list
        for record in x:
            curItem = self.myTree.item(record) # for each selected rows; get item 
            seen = set(self.deleteName) # avoid duplicate in to be deleted list
            item = curItem['values'][1] # Grab Name from tree view
            if item not in seen:
                seen.add(item)
                self.append_deleteName(item) # Append selected name to to be deleted list

            for name in self.deleteName:
                self.end.drop(self.end.loc[self.end['name']== name].index, inplace=True) # Drop names in database for Generate draft email

            print(self.end['name'])

            self.myTree.delete(record) # Delete rows from treeview

# ---------------------------------------------------------------------------- #
#                Draft an email including related staff details                #
# ---------------------------------------------------------------------------- #
    def reviewDraft(self):
        if len(self.myTree.get_children()) == 0:
            messagebox.showinfo(title="Error", message="List of staff is empty")
            return
        else:
            self.emailWindow()

    def emailWindow(self):

        newWindow = Toplevel(root)
        newWindow.title('Outlook Email Draft')
        newWindow.geometry('820x490') 
        newWindow.resizable(0,0)

        labelEntry = Frame(newWindow)
        labelEntry.grid(row =0 ,column = 0, sticky ='w')

        labelFrame = Frame(labelEntry)
        labelFrame.grid(row = 0, column = 1, sticky = 'w') 
        entryFrame = Frame(labelEntry)
        entryFrame.grid(row = 0, column = 2, sticky = 'w') 
        textFrame = Frame(newWindow)
        textFrame.grid(row = 2, column = 0, sticky = 'w')
        buttonFrame = Frame(labelEntry)
        buttonFrame.grid(row = 0, column = 0, sticky = 'w')

        Label(labelFrame, text = 'CC:').grid(row = 0, column = 0, padx = 5, pady = 5, sticky = 'w') 
        Label(labelFrame, text = 'Subject:').grid(row = 1, column = 0, padx = 5, pady = 5, sticky = 'w')

        self.emailTemp = Text(textFrame, state = 'normal', wrap = 'word', width = 100, height = 25)
        self.emailTemp.grid(row = 1, column = 0, padx = 5, pady=5)

        try:
            self.emailTemp.insert(END, self.emailBody)
        except AttributeError:
            newWindow.destroy()
            return

        self.emailTemp.config(state='disabled')

        self.ccEntry = Entry(entryFrame, textvariable = self.defCC ,width = 57)
        self.ccEntry.grid(row = 0, column = 1, padx =5, pady =5, sticky = 'e') 

        self.subjectEntry = Entry(entryFrame, width = 57, textvariable = self.emailSubject, state = 'normal')
        self.subjectEntry.grid(row = 1, column = 1, padx =5, pady =5, sticky = 'e')
        self.subjectEntry.delete(0, END)
        self.subjectEntry.insert(END, self.emailSubject)
        self.subjectEntry.config(state='disabled')


        self.editTemp = Button(buttonFrame, command = self.editTemplate, text ='Edit Template', width = 15)
        self.editTemp.grid(row = 0, column = 0, padx = 5, pady =5)

        self.saveDraft = Button(buttonFrame, command = self.toOutlook, text ='Save as Draft', width = 15)
        self.saveDraft.grid(row = 1, column = 0, padx = 5, pady =5)

    def editTemplate(self):
        self.subjectEntry.config(state = 'normal')
        self.emailTemp.config(state='normal')

    def toOutlook(self):
        self.end.loc[:,'endprobation'] = self.end.loc[:,'endprobation'].apply(pd.to_datetime)
        self.end.loc[:,'endprobation'] = self.end['endprobation'].dt.strftime('%d %B %Y')
# ----------------- Launch Outlook and write unique messages ----------------- #
        outlook = client.Dispatch("Outlook.Application")

        for info in range(len(self.rowLabels)):
            self.pfNumber = self.end.loc[self.end.index[info],'pers.no.']
            self.gender = self.end.loc[self.end.index[info],'gender']
            self.name = self.end.loc[self.end.index[info],'name']
            self.email = self.end.loc[self.end.index[info],'email']
            self.dueDate = self.end.loc[self.end.index[info], 'endprobation']

            message = outlook.CreateItem(0)
            message.To = self.email
            message.CC = self.ccEntry.get()

            self.emailSubject = self.subjectEntry.get()
            message.Subject = self.emailSubject.format(self.name, self.pfNumber)

            self.emailBody = self.emailTemp.get('1.0', END)
            message.Body = self.emailBody.format('Ms.' if self.gender == 'Female' else 'Mr.', self.name.split(" ")[0], self.dueDate)
            message.Display()

            with open('outlookSubject.txt', 'w') as f:
                    f.write(f"{self.emailSubject}")

            if self.emailBody == " " or "Dear" not in (self.emailBody) or r"{}" not in (self.emailBody):
                pass
            else:
                with open('outlookBody.txt', 'w') as f:
                    f.write(f"{self.emailBody}")

        self.end.loc[:,'endprobation'] = self.end.loc[:,'endprobation'].apply(pd.to_datetime)
        

        print("successfully draft all email")

    def entry(self):
        Label(self.wrapper, text = 'File Name:').grid(row = 0, column = 0, padx = 5, pady = 5, sticky = 'e')
        Label(self.wrapper, text = 'No. of Days:').grid(row = 1, column = 0, padx = 5, pady = 5)
        
        self.pathEntry = Entry(self.wrapper, textvariable = self.fileName, width = 57, state = 'disabled')
        self.pathEntry.grid(row = 0, column = 1, padx = 5, pady = 5)

        self.daysEntry = Entry(self.wrapper, textvariable = '', width = 25)
        self.daysEntry.grid(row = 1, column = 1, padx = 5, pady = 5, sticky = 'w')

    def button(self):
        self.Filter = Button(self.wrapper, command = lambda:[self.number(),self.clearTree(),self.showList()], text = 'Filter', width = 10)
        self.Filter.bind('<Return>',self.parse)
        self.Filter.grid(row = 1, column = 1, padx= 10, pady =5, sticky = 'e')

        self.deleteRow = Button(self.treeWrapper, command = self.removeMany,text ='Delete', width = 10)
        self.deleteRow.grid(row = 0, column = 0, padx = 10, pady =5)
        
        self.toMail = Button(self.treeWrapper, command = self.reviewDraft, text ='Export to Mail', width = 15)
        self.toMail.grid(row = 0, column = 1, padx = 10, pady =5)
# ---------------------------------------------------------------------------- #
#                     Press Filter Button using 'Enter' Key                    #
# ---------------------------------------------------------------------------- #
    def parse(self,event):
        self.number()
        self.showList()

if __name__ == "__main__":
    root = tk.Tk()
    Application(root)
    root.mainloop()


